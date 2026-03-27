"""
Multi-tab Excel manager using openpyxl.
Exports pipeline results to a structured Excel workbook.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from pylon.models import PipelineContext

_logger = logging.getLogger("excel")

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")


class ExcelManager:
    """Exports pipeline results to a multi-tab Excel workbook."""

    def __init__(self, output_dir: str | Path | None = None) -> None:
        self.output_dir = Path(output_dir) if output_dir else Path.cwd() / "output"

    def export(self, context: PipelineContext) -> str:
        """
        Export pipeline context to an Excel file.
        Creates tabs: Summary, Companies, Skills, Contacts, Outreach.
        Returns the file path.
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"pylon_{context.run_id[:8]}_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        wb = Workbook()

        # Tab 1: Summary
        ws = wb.active
        ws.title = "Summary"
        self._write_summary(ws, context)

        # Tab 2: Companies
        ws_companies = wb.create_sheet("Companies")
        self._write_companies(ws_companies, context)

        # Tab 3: Skills (if available)
        if context.skills:
            ws_skills = wb.create_sheet("Skills")
            self._write_skills(ws_skills, context)

        # Tab 4: Contacts (if available)
        if context.contacts:
            ws_contacts = wb.create_sheet("Contacts")
            self._write_contacts(ws_contacts, context)

        # Tab 5: Outreach (if available)
        if context.drafts:
            ws_outreach = wb.create_sheet("Outreach")
            self._write_outreach(ws_outreach, context)

        wb.save(str(filepath))
        _logger.info("Excel exported to %s", filepath)
        return str(filepath)

    def _write_header(self, ws: Any, headers: list[str]) -> None:
        """Write a formatted header row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT_WHITE
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    def _write_summary(self, ws: Any, context: PipelineContext) -> None:
        """Write the Summary tab."""
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 60

        rows = [
            ("Query", context.query),
            ("Run ID", context.run_id),
            ("Date", context.created_at.strftime("%Y-%m-%d %H:%M UTC")),
            ("Companies Found", len(context.candidates)),
            ("Profiles Researched", len(context.profiles)),
            ("Skills Analyzed", len(context.skills)),
            ("Contacts Found", len(context.contacts)),
            ("Drafts Created", len(context.drafts)),
        ]
        for row_idx, (label, value) in enumerate(rows, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=str(value))

    def _write_companies(self, ws: Any, context: PipelineContext) -> None:
        """Write the Companies tab."""
        headers = ["Name", "Domain", "Relevance", "Website", "Confidence",
                    "Funding", "HQ", "Employees", "Hiring Signals"]
        self._write_header(ws, headers)

        for row_idx, c in enumerate(context.candidates, 2):
            profile = next((p for p in context.profiles if p.company_name == c.name), None)
            ws.cell(row=row_idx, column=1, value=c.name)
            ws.cell(row=row_idx, column=2, value=c.domain.value)
            ws.cell(row=row_idx, column=3, value=c.relevance_reason)
            ws.cell(row=row_idx, column=4, value=c.website)
            ws.cell(row=row_idx, column=5, value=c.confidence)
            if profile:
                ws.cell(row=row_idx, column=6, value=profile.funding_stage.value)
                ws.cell(row=row_idx, column=7, value=profile.headquarters)
                ws.cell(row=row_idx, column=8, value=profile.employee_count)
                ws.cell(row=row_idx, column=9, value=", ".join(profile.hiring_signals))

        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _write_skills(self, ws: Any, context: PipelineContext) -> None:
        """Write the Skills tab."""
        headers = ["Company", "Tools", "ML Frameworks", "Cloud", "Alignment",
                    "Skills to Learn", "Gap Analysis"]
        self._write_header(ws, headers)

        for row_idx, s in enumerate(context.skills, 2):
            ws.cell(row=row_idx, column=1, value=s.company_name)
            ws.cell(row=row_idx, column=2, value=", ".join(s.tools_used))
            ws.cell(row=row_idx, column=3, value=", ".join(s.ml_frameworks))
            ws.cell(row=row_idx, column=4, value=s.cloud_platform)
            ws.cell(row=row_idx, column=5, value=s.alignment_score)
            ws.cell(row=row_idx, column=6, value=", ".join(s.skills_to_learn))
            ws.cell(row=row_idx, column=7, value=s.gap_analysis)

        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 22

    def _write_contacts(self, ws: Any, context: PipelineContext) -> None:
        """Write the Contacts tab."""
        headers = ["Company", "Name", "Title", "Email", "LinkedIn", "Notes", "Confidence"]
        self._write_header(ws, headers)

        for row_idx, ct in enumerate(context.contacts, 2):
            ws.cell(row=row_idx, column=1, value=ct.company_name)
            ws.cell(row=row_idx, column=2, value=ct.name)
            ws.cell(row=row_idx, column=3, value=ct.title)
            ws.cell(row=row_idx, column=4, value=ct.email)
            ws.cell(row=row_idx, column=5, value=ct.linkedin_url)
            ws.cell(row=row_idx, column=6, value=ct.notes)
            ws.cell(row=row_idx, column=7, value=ct.confidence)

        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 22

    def _write_outreach(self, ws: Any, context: PipelineContext) -> None:
        """Write the Outreach tab."""
        headers = ["Company", "Contact", "Subject", "Body", "Status", "Template"]
        self._write_header(ws, headers)

        for row_idx, d in enumerate(context.drafts, 2):
            ws.cell(row=row_idx, column=1, value=d.company_name)
            ws.cell(row=row_idx, column=2, value=d.contact_name)
            ws.cell(row=row_idx, column=3, value=d.subject)
            ws.cell(row=row_idx, column=4, value=d.body)
            ws.cell(row=row_idx, column=5, value=d.status.value)
            ws.cell(row=row_idx, column=6, value=d.template_used)

        ws.column_dimensions["D"].width = 60
        for col in [1, 2, 3, 5, 6]:
            ws.column_dimensions[chr(64 + col)].width = 20
