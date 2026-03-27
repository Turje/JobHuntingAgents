"""Tests for src/pylon/excel.py — ExcelManager."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from pylon.excel import ExcelManager
from pylon.models import (
    CompanyCandidate,
    CompanyProfile,
    ContactInfo,
    FundingStage,
    IndustryDomain,
    OutreachDraft,
    PipelineContext,
    SkillsAnalysis,
)


@pytest.fixture
def context():
    ctx = PipelineContext.new("find football analytics companies")
    ctx.candidates = [
        CompanyCandidate(
            name="StatsBomb",
            domain=IndustryDomain.SPORTS_TECH,
            relevance_reason="Football analytics",
            website="https://statsbomb.com",
            confidence=0.9,
        ),
    ]
    ctx.profiles = [
        CompanyProfile(
            company_name="StatsBomb",
            funding_stage=FundingStage.SERIES_B,
            headquarters="Bath, UK",
            employee_count="50-100",
            hiring_signals=["ML Engineer role posted"],
        ),
    ]
    ctx.skills = [
        SkillsAnalysis(
            company_name="StatsBomb",
            tools_used=["Python", "PostgreSQL"],
            ml_frameworks=["PyTorch"],
            cloud_platform="AWS",
            alignment_score=0.75,
        ),
    ]
    ctx.contacts = [
        ContactInfo(company_name="StatsBomb", name="Ted Knutson", title="CEO"),
    ]
    ctx.drafts = [
        OutreachDraft(company_name="StatsBomb", contact_name="Ted", subject="Hello"),
    ]
    return ctx


class TestExcelManager:
    def test_export_creates_file(self, tmp_path, context):
        mgr = ExcelManager(output_dir=tmp_path)
        path = mgr.export(context)
        assert Path(path).exists()
        assert path.endswith(".xlsx")

    def test_summary_tab(self, tmp_path, context):
        mgr = ExcelManager(output_dir=tmp_path)
        path = mgr.export(context)
        wb = load_workbook(path)
        assert "Summary" in wb.sheetnames
        ws = wb["Summary"]
        assert ws.cell(row=1, column=1).value == "Query"
        assert ws.cell(row=1, column=2).value == context.query

    def test_companies_tab(self, tmp_path, context):
        mgr = ExcelManager(output_dir=tmp_path)
        path = mgr.export(context)
        wb = load_workbook(path)
        assert "Companies" in wb.sheetnames
        ws = wb["Companies"]
        assert ws.cell(row=1, column=1).value == "Name"
        assert ws.cell(row=2, column=1).value == "StatsBomb"

    def test_skills_tab(self, tmp_path, context):
        mgr = ExcelManager(output_dir=tmp_path)
        path = mgr.export(context)
        wb = load_workbook(path)
        assert "Skills" in wb.sheetnames

    def test_contacts_tab(self, tmp_path, context):
        mgr = ExcelManager(output_dir=tmp_path)
        path = mgr.export(context)
        wb = load_workbook(path)
        assert "Contacts" in wb.sheetnames
        ws = wb["Contacts"]
        assert ws.cell(row=2, column=2).value == "Ted Knutson"

    def test_outreach_tab(self, tmp_path, context):
        mgr = ExcelManager(output_dir=tmp_path)
        path = mgr.export(context)
        wb = load_workbook(path)
        assert "Outreach" in wb.sheetnames

    def test_no_optional_tabs_when_empty(self, tmp_path, context):
        context.skills = []
        context.contacts = []
        context.drafts = []
        mgr = ExcelManager(output_dir=tmp_path)
        path = mgr.export(context)
        wb = load_workbook(path)
        assert "Skills" not in wb.sheetnames
        assert "Contacts" not in wb.sheetnames
        assert "Outreach" not in wb.sheetnames

    def test_output_dir_created(self, tmp_path, context):
        output = tmp_path / "nested" / "output"
        mgr = ExcelManager(output_dir=output)
        path = mgr.export(context)
        assert Path(path).exists()
